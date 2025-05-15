import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import datetime
import json
import io
import base64
from io import BytesIO, StringIO
import os
import time
import re

from database import (
    get_products, get_product, add_product, update_product, delete_product, 
    get_price_history, get_settings, update_settings, get_suggested_prices,
    add_suggested_price, update_suggested_price, delete_suggested_price,
    get_latest_prices, export_prices_to_json, export_prices_to_csv
)
from scraper import (
    scrape_all_products, test_scrape, get_scheduler_status, 
    start_scheduler, stop_scheduler, run_scraper_now
)
from analyzers import (
    get_price_analysis, get_bulk_analysis,
    create_price_history_chart, create_price_statistics_table,
    create_price_comparison_gauge_chart, create_price_trend_forecast,
    create_competitor_price_matrix, create_price_difference_chart
)

# Monitor Products Page
def monitor_products_page():
    st.title("📊 Monitor Products")
    
    # Get products from database
    products_df = get_products()
    
    if products_df.empty:
        st.warning("No products found. Please add products first.")
        return
    
    # Create tabs
    overview_tab, details_tab = st.tabs(["Products Overview", "Product Details"])
    
    with overview_tab:
        st.subheader("Products Overview")
        
        # Add status indicator
        status_cols = st.columns(3)
        
        with status_cols[0]:
            # Count total products
            st.metric("Total Products", len(products_df))
        
        with status_cols[1]:
            # Last update time
            if 'last_checked' in products_df.columns:
                last_update = pd.to_datetime(products_df['last_checked']).max()
                if pd.notna(last_update):
                    last_update_str = last_update.strftime('%Y-%m-%d %H:%M')
                    st.metric("Last Updated", last_update_str)
                else:
                    st.metric("Last Updated", "Never")
            else:
                st.metric("Last Updated", "Never")
        
        with status_cols[2]:
            # Scheduler status
            scheduler_status = get_scheduler_status()
            status_text = "Active" if scheduler_status["running"] else "Inactive"
            st.metric("Auto-Scraper", status_text)
        
        # Create a run now button
        if st.button("Run Scraper Now", type="primary"):
            with st.spinner("Scraping products..."):
                results = run_scraper_now()
                st.success(f"Scraping completed: {results.get('scraped', 0)} products scraped, {results.get('errors', 0)} errors")
                # Refresh the page to update data
                st.rerun()
        
        # Add product list with status
        st.markdown("### Product Status")
        
        # Function to format status
        def format_status(status):
            if status:
                return "✅ Active"
            else:
                return "❌ Not scraped yet"
        
        # Create a DataFrame with essential columns for display
        if not products_df.empty:
            display_df = products_df[['id', 'name', 'last_checked', 'current_price']].copy()
            
            # Add status column
            display_df['status'] = display_df['last_checked'].notna()
            display_df['status'] = display_df['status'].apply(format_status)
            
            # Rename columns
            display_df = display_df.rename(columns={
                'id': 'ID',
                'name': 'Product',
                'last_checked': 'Last Checked',
                'current_price': 'Current Price',
                'status': 'Status'
            })
            
            # Format datetime columns
            if 'Last Checked' in display_df.columns:
                display_df['Last Checked'] = pd.to_datetime(display_df['Last Checked']).dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # Format price columns
            if 'Current Price' in display_df.columns:
                display_df['Current Price'] = display_df['Current Price'].apply(lambda x: f"€{x:.2f}" if pd.notna(x) else "N/A")
            
            # Display the DataFrame
            st.dataframe(display_df, use_container_width=True)
    
    with details_tab:
        st.subheader("Product Details")
        
        # Create a selectbox for product selection
        product_options = [(row['id'], row['name']) for _, row in products_df.iterrows()]
        selected_product = st.selectbox(
            "Select Product",
            options=product_options,
            format_func=lambda x: f"{x[1]}"
        )
        
        if selected_product:
            product_id = selected_product[0]
            
            # Get product details
            product = get_product(product_id)
            
            if product:
                st.markdown(f"### {product['name']}")
                
                # Create tabs for different views
                price_tab, config_tab = st.tabs(["Price History", "Product Configuration"])
                
                with price_tab:
                    # Get price history for the selected product
                    price_history = get_price_history(product_id)
                    
                    if not price_history.empty:
                        st.success(f"Found {len(price_history)} price records")
                        
                        # Display price history chart
                        view_mode = st.radio(
                            "Chart Type",
                            options=["line", "area", "bar", "candlestick"],
                            index=0,
                            horizontal=True
                        )
                        
                        fig = create_price_history_chart(price_history, product['name'], view_mode)
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Display price statistics
                        st.markdown("### Price Statistics")
                        stats_df = create_price_statistics_table(price_history)
                        if not stats_df.empty:
                            st.dataframe(stats_df, use_container_width=True)
                        
                        # Display competitor comparison gauge
                        st.markdown("### Market Position")
                        gauge_fig = create_price_comparison_gauge_chart(price_history, product['name'])
                        st.plotly_chart(gauge_fig, use_container_width=True)
                        
                        # Display price vs competitors chart
                        st.markdown("### Price Difference Analysis")
                        diff_fig = create_price_difference_chart(price_history, product['name'])
                        st.plotly_chart(diff_fig, use_container_width=True)
                        
                        # Display competitor price matrix
                        st.markdown("### Competitor Price Matrix")
                        matrix_fig = create_competitor_price_matrix(price_history, product['name'])
                        st.plotly_chart(matrix_fig, use_container_width=True)
                    else:
                        st.warning("No price history found for this product. Please run the scraper to collect data.")
                
                with config_tab:
                    st.markdown("### Product Configuration Details")
                    
                    # Create columns
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("#### Basic Information")
                        st.markdown(f"**ID:** {product['id']}")
                        st.markdown(f"**Name:** {product['name']}")
                        st.markdown(f"**URL:** {product['our_url']}")
                        st.markdown(f"**Price Selector:** `{product['our_price_selector']}`")
                        
                        if 'our_name_selector' in product and product['our_name_selector']:
                            st.markdown(f"**Name Selector:** `{product['our_name_selector']}`")
                        
                        # Price thresholds
                        st.markdown("#### Price Thresholds")
                        min_threshold = product.get('min_price_threshold', "Not set (using global)")
                        max_threshold = product.get('max_price_threshold', "Not set (using global)")
                        st.markdown(f"**Min Threshold:** {min_threshold}€")
                        st.markdown(f"**Max Threshold:** {max_threshold}€")
                        
                        # Test selectors for our product
                        st.markdown("#### Test Our Selectors")
                        test_cols = st.columns(2)
                        with test_cols[0]:
                            if st.button("Test Our Name Selector"):
                                with st.spinner("Testing..."):
                                    name_selector = product.get('our_name_selector')
                                    if name_selector:
                                        result = test_scrape(product['our_url'], None, name_selector)
                                        if result and 'name' in result:
                                            st.success(f"Found name: {result['name']}")
                                        else:
                                            st.error(f"Failed to find name: {result.get('error', 'Unknown error')}")
                                    else:
                                        st.warning("No name selector defined")
                        
                        with test_cols[1]:
                            if st.button("Test Our Price Selector"):
                                with st.spinner("Testing..."):
                                    result = test_scrape(product['our_url'], product['our_price_selector'])
                                    if result and 'price' in result:
                                        st.success(f"Found price: {result['price']}")
                                    else:
                                        st.error(f"Failed to find price: {result.get('error', 'Unknown error')}")
                    
                    with col2:
                        st.markdown("#### Competitor Information")
                        
                        # Get competitor URLs
                        competitor_urls = product.get('competitor_urls', {})
                        competitor_selectors = product.get('competitor_selectors', {})
                        
                        if competitor_urls:
                            st.markdown(f"Found {len(competitor_urls)} competitors")
                            
                            # Display each competitor
                            for idx, (comp_id, comp_url) in enumerate(competitor_urls.items()):
                                
                                st.markdown(f"**Competitor {idx+1}**")
                                st.markdown(f"URL: {comp_url}")
                                
                                # Handle different formats of competitor selectors
                                name_selector = "Not set"
                                price_selector = "Not set"
                                
                                # Try to get selectors from different formats
                                if isinstance(competitor_selectors, dict):
                                    # Check if comp_id is in the dict
                                    if comp_id in competitor_selectors:
                                        selectors = competitor_selectors[comp_id]
                                        if isinstance(selectors, dict):
                                            name_selector = selectors.get('name', 'Not set')
                                            price_selector = selectors.get('price', 'Not set')
                                        elif isinstance(selectors, list) and len(selectors) >= 2:
                                            name_selector = selectors[0]
                                            price_selector = selectors[1]
                                    # Try index-based access if numeric
                                    elif comp_id.isdigit() and int(comp_id) < len(competitor_selectors):
                                        idx_selectors = list(competitor_selectors.values())[int(comp_id)]
                                        if isinstance(idx_selectors, dict):
                                            name_selector = idx_selectors.get('name', 'Not set')
                                            price_selector = idx_selectors.get('price', 'Not set')
                                # List of lists format [[name1, price1], [name2, price2]]
                                elif isinstance(competitor_selectors, list) and idx < len(competitor_selectors):
                                    idx_selectors = competitor_selectors[idx]
                                    if isinstance(idx_selectors, list) and len(idx_selectors) >= 2:
                                        name_selector = idx_selectors[0]
                                        price_selector = idx_selectors[1]
                                    elif isinstance(idx_selectors, dict):
                                        name_selector = idx_selectors.get('name', 'Not set')
                                        price_selector = idx_selectors.get('price', 'Not set')
                                        
                                st.markdown(f"Name Selector: `{name_selector}`")
                                st.markdown(f"Price Selector: `{price_selector}`")
                                
                                # Add test buttons
                                test_cols = st.columns(2)
                                with test_cols[0]:
                                    if st.button(f"Test Name {idx+1}", key=f"test_name_{comp_id}"):
                                        from scraper import test_scrape
                                        result = test_scrape(comp_url, None, name_selector)
                                        if result and result.get('name'):
                                            st.success(f"Found: {result['name']}")
                                        else:
                                            st.error("Failed to find element with selector")
                                        
                                with test_cols[1]:
                                    if st.button(f"Test Price {idx+1}", key=f"test_price_{comp_id}"):
                                        from scraper import test_scrape
                                        result = test_scrape(comp_url, price_selector)
                                        if result and result.get('price') is not None:
                                            st.success(f"Found price: {result['price']}")
                                        else:
                                            st.error("Failed to find price with selector")
            else:
                st.error("Could not retrieve detailed product information.")
            
        # Display the price history chart
        # ... (rest of the price history display)

# Add Product Page
def add_product_page():
    st.title("➕ Add Product")
    
    # Create tabs for manual and batch adding
    manual_tab, batch_tab = st.tabs(["Add Single Product", "Batch Import"])
    
    with manual_tab:
        st.markdown("""
        ### Add a single product
        Enter the details of the product you want to monitor.
        """)
        
        # Create a form for adding a product
        with st.form("add_product_form"):
            # Product details
            name = st.text_input("Product Name", help="A descriptive name for the product")
            our_url = st.text_input("Product URL", help="The URL of the product on your store")
            our_name_selector = st.text_input("Product Name Selector (optional)", help="CSS selector for the product name")
            our_price_selector = st.text_input("Product Price Selector", help="CSS selector for the product price")
            
            # Price thresholds
            st.markdown("### Price Thresholds (Optional)")
            st.markdown("Set price variation thresholds in absolute EUR amounts.")
            
            threshold_cols = st.columns(2)
            with threshold_cols[0]:
                min_price_threshold = st.number_input("Min Price Threshold (€)", 
                                                    help="Minimum allowed price variation in EUR (usually negative)",
                                                    min_value=-1000.0, max_value=0.0, value=-5.0, step=1.0)
            with threshold_cols[1]:
                max_price_threshold = st.number_input("Max Price Threshold (€)",
                                                    help="Maximum allowed price variation in EUR (usually positive)",
                                                    min_value=0.0, max_value=1000.0, value=15.0, step=1.0)
            
            # Competitor details
            st.markdown("### Competitor URLs (Optional)")
            st.markdown("Add up to 5 competitors to compare prices with.")
            
            competitor_urls = {}
            competitor_selectors = {}
            
            for i in range(5):
                st.markdown(f"#### Competitor {i+1}")
                comp_cols = st.columns(2)
                
                with comp_cols[0]:
                    comp_url = st.text_input(f"Competitor {i+1} URL", key=f"comp_url_{i}")
                
                if comp_url:
                    competitor_urls[str(i)] = comp_url
                    
                    with comp_cols[1]:
                        comp_name_selector = st.text_input(f"Competitor {i+1} Name Selector", key=f"comp_name_{i}")
                        comp_price_selector = st.text_input(f"Competitor {i+1} Price Selector", key=f"comp_price_{i}")
                    
                    if comp_price_selector:
                        competitor_selectors[str(i)] = {
                            "name": comp_name_selector,
                            "price": comp_price_selector
                        }
            
            # Submit button
            submit = st.form_submit_button("Add Product")
            
            if submit:
                try:
                    # Add the product to the database
                    product_id = add_product(
                        name=name,
                        our_url=our_url,
                        our_name_selector=our_name_selector,
                        our_price_selector=our_price_selector,
                        competitor_urls=competitor_urls,
                        competitor_selectors=competitor_selectors,
                        min_price_threshold=min_price_threshold,
                        max_price_threshold=max_price_threshold
                    )
                    
                    st.success(f"Product '{name}' added successfully with ID {product_id}!")
                    
                    # Ask if user wants to test selectors
                    st.info("Product added. You can now test the selectors.")
                except Exception as e:
                    st.error(f"Error adding product: {str(e)}")
        
        # Add test button outside the form to prevent form submission
        st.markdown("### Test Selectors")
        test_cols = st.columns(2)
        
        with test_cols[0]:
            if st.button("Test Our Price Selector"):
                if our_url and our_price_selector:
                    with st.spinner("Testing price selector..."):
                        result = test_scrape(our_url, our_price_selector)
                        if result and 'price' in result and result['price'] is not None:
                            st.success(f"Successfully found price: {result['price']}")
                        else:
                            st.error(f"Failed to find price: {result.get('error', 'Unknown error')}")
                else:
                    st.warning("Please enter URL and price selector first")
        
        with test_cols[1]:
            if st.button("Test Our Name Selector"):
                if our_url and our_name_selector:
                    with st.spinner("Testing name selector..."):
                        result = test_scrape(our_url, None, our_name_selector)
                        if result and 'name' in result:
                            st.success(f"Successfully found name: {result['name']}")
                        else:
                            st.error(f"Failed to find name: {result.get('error', 'Unknown error')}")
                else:
                    st.warning("Please enter URL and name selector first")
    
    with batch_tab:
        st.markdown("""
        ### Import products from Excel
        Upload an Excel file with product details for batch import. 
        """)
        
        # Add a button to download the template
        # from utils.excel_importer import generate_template_excel
        if st.button("Download Template Excel"):
            # Generate template Excel
            import pandas as pd
            from openpyxl import Workbook
            from io import BytesIO
            
            # Create a workbook with a template
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Add headers
            headers = [
                "product_name", "our_url", "our_name_selector", "our_price_selector",
                "competitor1_url", "competitor1_name_selector", "competitor1_price_selector",
                "competitor2_url", "competitor2_name_selector", "competitor2_price_selector",
                "competitor3_url", "competitor3_name_selector", "competitor3_price_selector",
                "min_price_threshold", "max_price_threshold"
            ]
            
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Add a sample row
            sample_data = [
                "Sample Product", "https://example.com/product", "#product-name", ".product-price",
                "https://competitor1.com/product", "#product-title", ".price",
                "https://competitor2.com/product", ".product-name", "#price",
                "", "", "",
                "-5", "15"
            ]
            
            for col_num, value in enumerate(sample_data, 1):
                ws.cell(row=2, column=col_num, value=value)
            
            # Save to a BytesIO object
            excel_data = BytesIO()
            wb.save(excel_data)
            excel_data.seek(0)
            
            # Create a download link
            b64 = base64.b64encode(excel_data.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="product_import_template.xlsx">Download Excel Template</a>'
            st.markdown(href, unsafe_allow_html=True)
        
        # Upload Excel file
        uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])
        
        if uploaded_file:
            try:
                # Process the Excel file
                df = pd.read_excel(uploaded_file)
                
                if 'product_name' not in df.columns or 'our_url' not in df.columns or 'our_price_selector' not in df.columns:
                    st.error("Excel file must contain at least 'product_name', 'our_url', and 'our_price_selector' columns.")
                else:
                    st.success(f"Found {len(df)} products in the Excel file.")
                    st.dataframe(df.head(10))
                    
                    # Process and import
                    if st.button("Import Products"):
                        imported = 0
                        errors = 0
                        
                        for _, row in df.iterrows():
                            try:
                                # Extract basic product data
                                name = row['product_name']
                                our_url = row['our_url']
                                our_name_selector = row.get('our_name_selector', '')
                                our_price_selector = row['our_price_selector']
                                
                                # Extract price thresholds
                                min_price_threshold = float(row['min_price_threshold']) if pd.notna(row.get('min_price_threshold')) else None
                                max_price_threshold = float(row['max_price_threshold']) if pd.notna(row.get('max_price_threshold')) else None
                                
                                # Extract competitor data
                                competitor_urls = {}
                                competitor_selectors = {}
                                
                                # Look for competitor columns
                                for i in range(1, 6):  # Up to 5 competitors
                                    comp_url_col = f"competitor{i}_url"
                                    comp_name_sel_col = f"competitor{i}_name_selector"
                                    comp_price_sel_col = f"competitor{i}_price_selector"
                                    
                                    if comp_url_col in row and pd.notna(row[comp_url_col]) and row[comp_url_col]:
                                        comp_url = row[comp_url_col]
                                        comp_name_sel = row.get(comp_name_sel_col, '')
                                        comp_price_sel = row.get(comp_price_sel_col, '')
                                        
                                        if comp_url:
                                            competitor_urls[str(i-1)] = comp_url
                                            
                                            if comp_price_sel:
                                                competitor_selectors[str(i-1)] = {
                                                    "name": comp_name_sel,
                                                    "price": comp_price_sel
                                                }
                                
                                # Add the product
                                product_id = add_product(
                                    name=name,
                                    our_url=our_url,
                                    our_name_selector=our_name_selector,
                                    our_price_selector=our_price_selector,
                                    competitor_urls=competitor_urls,
                                    competitor_selectors=competitor_selectors,
                                    min_price_threshold=min_price_threshold,
                                    max_price_threshold=max_price_threshold
                                )
                                
                                imported += 1
                            except Exception as e:
                                errors += 1
                                st.error(f"Error importing row for '{row.get('product_name', 'Unknown')}': {str(e)}")
                        
                        if imported > 0:
                            st.success(f"Successfully imported {imported} products. {errors} errors occurred.")
                        else:
                            st.error(f"Failed to import any products. {errors} errors occurred.")
            except Exception as e:
                st.error(f"Error processing Excel file: {str(e)}")

# Price Analysis Page
def price_analysis_page():
    st.title("🔍 Price Analysis")
    
    # Get products from database
    products_df = get_products()
    
    if products_df.empty:
        st.warning("No products found. Please add products first.")
        return
    
    # Create a selectbox for product selection
    product_options = [(row['id'], row['name']) for _, row in products_df.iterrows()]
    selected_product = st.selectbox(
        "Select Product",
        options=product_options,
        format_func=lambda x: f"{x[1]}"
    )
    
    if selected_product:
        product_id = selected_product[0]
        product_name = selected_product[1]
        
        # Analysis settings
        st.markdown("### Analysis Settings")
        
        # Analysis period
        settings = get_settings()
        default_days = int(settings.get("analysis_period", 7))
        
        analysis_days = st.slider(
            "Analysis Period (days)",
            min_value=1,
            max_value=60,
            value=default_days,
            step=1,
            help="Number of days of price history to include in the analysis"
        )
        
        # Run analysis button
        if st.button("Run AI Analysis", type="primary"):
            with st.spinner("Analyzing price data..."):
                # Get price analysis
                analysis = get_price_analysis(product_id, analysis_days)
                
                if 'error' in analysis:
                    st.error(f"Analysis error: {analysis['error']}")
                else:
                    # Display analysis results
                    st.markdown(f"### AI Price Analysis for {product_name}")
                    
                    # Create columns for key metrics
                    metric_cols = st.columns(3)
                    
                    with metric_cols[0]:
                        current_price = analysis.get('current_price', 0)
                        st.metric("Current Price", f"€{current_price:.2f}")
                    
                    with metric_cols[1]:
                        avg_competitor = analysis.get('average_competitor_price', 0)
                        diff = current_price - avg_competitor if avg_competitor > 0 else 0
                        diff_pct = (diff / avg_competitor * 100) if avg_competitor > 0 else 0
                        st.metric("Avg. Competitor Price", f"€{avg_competitor:.2f}", delta=f"{diff_pct:.1f}%")
                    
                    with metric_cols[2]:
                        # Market position
                        position = analysis.get('price_position', 'unknown').title()
                        st.metric("Market Position", position)
                    
                    # Show price suggestion
                    if 'suggested_price' in analysis and analysis['suggested_price'] is not None:
                        suggested_price = analysis['suggested_price']
                        price_diff = suggested_price - current_price
                        price_diff_pct = (price_diff / current_price * 100) if current_price > 0 else 0
                        
                        suggestion_container = st.container(border=True)
                        with suggestion_container:
                            st.markdown("### Price Suggestion")
                            sugg_cols = st.columns(2)
                            
                            with sugg_cols[0]:
                                st.metric(
                                    "Suggested Price", 
                                    f"€{suggested_price:.2f}", 
                                    delta=f"{price_diff:.2f} ({price_diff_pct:.1f}%)"
                                )
                            
                            with sugg_cols[1]:
                                if st.button("Apply Suggestion", key="apply_suggestion"):
                                    with st.spinner("Applying suggestion..."):
                                        # Add the suggested price
                                        suggestion_id = add_suggested_price(
                                            product_id=product_id,
                                            suggested_price=suggested_price,
                                            source='ai',
                                            notes=analysis.get('short_recommendation', 'AI suggestion')
                                        )
                                        
                                        st.success(f"Suggestion saved! Go to 'Price Management' to apply it.")
                    
                    # Analysis details
                    with st.expander("AI Analysis Details", expanded=True):
                        # Display the analysis sections if available
                        if 'market_analysis' in analysis and analysis['market_analysis']:
                            st.markdown("#### Market Position Analysis")
                            st.write(analysis['market_analysis'])
                        
                        if 'trend_analysis' in analysis and analysis['trend_analysis']:
                            st.markdown("#### Price Trend Analysis")
                            st.write(analysis['trend_analysis'])
                        
                        if 'competitor_analysis' in analysis and analysis['competitor_analysis']:
                            st.markdown("#### Competitor Analysis")
                            st.write(analysis['competitor_analysis'])
                        
                        if 'recommendation' in analysis and analysis['recommendation']:
                            st.markdown("#### Recommendations")
                            st.write(analysis['recommendation'])
                        
                        if 'reasoning' in analysis and analysis['reasoning']:
                            st.markdown("#### Reasoning")
                            st.write(analysis['reasoning'])
                    
                    # Price visualizations
                    st.markdown("### Price Comparison")
                    
                    # Get price history for visualization
                    price_history = get_price_history(product_id, analysis_days)
                    
                    if not price_history.empty:
                        # Create the price history chart
                        fig = create_price_history_chart(price_history, product_name)
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Create price trend forecast
                        st.markdown("### Price Trend Forecast")
                        forecast_fig = create_price_trend_forecast(price_history, product_name)
                        st.plotly_chart(forecast_fig, use_container_width=True)
                    else:
                        st.warning("No price history data available for visualization.")

# Price Management Page
def price_management_page():
    st.title("💰 Price Management")
    
    # Create tabs for different views
    suggestions_tab, export_tab = st.tabs(["Price Suggestions", "Export Prices"])
    
    with suggestions_tab:
        st.markdown("""
        ### Price Suggestions
        View and manage AI-suggested prices and manually set prices.
        """)
        
        # Get price suggestions
        suggestions_df = get_suggested_prices()
        
        if suggestions_df.empty:
            st.info("No price suggestions available. Run AI analysis to get suggestions.")
        else:
            # Add buttons for bulk actions
            action_cols = st.columns(3)
            
            with action_cols[0]:
                apply_all = st.button("Apply All Suggestions", type="primary")
                
                if apply_all:
                    with st.spinner("Applying all suggestions..."):
                        for idx, row in suggestions_df.iterrows():
                            suggestion_id = row['id']
                            update_suggested_price(suggestion_id, is_applied=True)
                        
                        st.success("All suggestions applied!")
                        # Refresh the page
                        st.rerun()
            
            with action_cols[1]:
                if st.button("Get New Suggestions for All"):
                    st.info("Redirecting to Multi-Product Analysis page...")
                    # Navigate to the bulk analysis page
                    # In Streamlit we can't directly navigate, so we just show a message
            
            # Display the suggestions table
            st.markdown("#### Current Suggestions")
            
            # Create a clean version of the DataFrame for display
            display_df = suggestions_df[['id', 'product_name', 'current_price', 'suggested_price', 
                                        'manual_price', 'source', 'timestamp', 'notes']].copy()
            
            # Format price columns
            for col in ['current_price', 'suggested_price', 'manual_price']:
                if col in display_df.columns:
                    display_df[col] = display_df[col].apply(
                        lambda x: f"€{x:.2f}" if pd.notna(x) else "N/A"
                    )
            
            # Format timestamp
            if 'timestamp' in display_df.columns:
                display_df['timestamp'] = pd.to_datetime(display_df['timestamp']).dt.strftime('%Y-%m-%d %H:%M')
            
            # Rename columns
            display_df = display_df.rename(columns={
                'id': 'ID',
                'product_name': 'Product',
                'current_price': 'Current Price',
                'suggested_price': 'Suggested Price',
                'manual_price': 'Manual Price',
                'source': 'Source',
                'timestamp': 'Created',
                'notes': 'Notes'
            })
            
            st.dataframe(display_df, use_container_width=True)
            
            # Individual suggestion management
            st.markdown("#### Manage Individual Suggestion")
            
            # Create a selectbox for suggestion selection
            suggestion_options = [(row['id'], f"{row['product_name']} (ID: {row['id']})") for _, row in suggestions_df.iterrows()]
            selected_suggestion = st.selectbox(
                "Select Suggestion",
                options=suggestion_options,
                format_func=lambda x: f"{x[1]}"
            )
            
            if selected_suggestion:
                suggestion_id = selected_suggestion[0]
                
                # Get the suggestion details
                suggestion = suggestions_df[suggestions_df['id'] == suggestion_id].iloc[0]
                
                # Display suggestion details
                st.markdown(f"**Product:** {suggestion['product_name']}")
                st.markdown(f"**Current Price:** €{suggestion['current_price']:.2f}")
                
                # Calculate price differences
                suggested_price = suggestion['suggested_price'] if pd.notna(suggestion['suggested_price']) else None
                if suggested_price:
                    price_diff = suggested_price - suggestion['current_price']
                    price_diff_pct = (price_diff / suggestion['current_price'] * 100) if suggestion['current_price'] > 0 else 0
                    st.markdown(f"**Suggested Price:** €{suggested_price:.2f} ({price_diff_pct:.1f}%)")
                
                # Create a form for updating the suggestion
                with st.form(key=f"update_suggestion_{suggestion_id}"):
                    # Manual price input
                    manual_price = st.number_input(
                        "Manual Price Override (€)",
                        min_value=0.01,
                        value=float(suggested_price) if suggested_price else None,
                        format="%.2f",
                        help="Enter a manual price to override the AI suggestion"
                    )
                    
                    # Notes
                    notes = st.text_area("Notes", value=suggestion['notes'] if pd.notna(suggestion['notes']) else "")
                    
                    # Action buttons
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        apply_button = st.form_submit_button("Apply Suggestion")
                    
                    with col2:
                        update_button = st.form_submit_button("Update Manual Price")
                    
                    with col3:
                        delete_button = st.form_submit_button("Delete Suggestion")
                    
                    if apply_button:
                        # Update the suggestion as applied
                        update_suggested_price(
                            suggestion_id=suggestion_id,
                            is_applied=True,
                            notes=notes
                        )
                        st.success("Suggestion applied!")
                        # Refresh the page
                        st.rerun()
                    
                    if update_button:
                        # Update the suggestion with manual price
                        update_suggested_price(
                            suggestion_id=suggestion_id,
                            manual_price=manual_price,
                            notes=notes
                        )
                        st.success("Manual price updated!")
                        # Refresh the page
                        st.rerun()
                    
                    if delete_button:
                        # Delete the suggestion
                        delete_suggested_price(suggestion_id)
                        st.success("Suggestion deleted!")
                        # Refresh the page
                        st.rerun()
    
    with export_tab:
        st.markdown("""
        ### Export Prices
        Export current and suggested prices in JSON or CSV format for integration with external systems.
        """)
        
        # Get the latest prices
        latest_prices = get_latest_prices()
        
        if latest_prices.empty:
            st.warning("No price data available to export.")
        else:
            # Display a preview of the data
            st.markdown("#### Data Preview")
            
            # Create a clean DataFrame for display
            preview_df = latest_prices[['id', 'name', 'current_price', 'final_suggested_price']].copy()
            
            # Format price columns
            for col in ['current_price', 'final_suggested_price']:
                if col in preview_df.columns:
                    preview_df[col] = preview_df[col].apply(
                        lambda x: f"€{x:.2f}" if pd.notna(x) else "N/A"
                    )
            
            # Rename columns
            preview_df = preview_df.rename(columns={
                'id': 'ID',
                'name': 'Product',
                'current_price': 'Current Price',
                'final_suggested_price': 'Suggested Price'
            })
            
            st.dataframe(preview_df, use_container_width=True)
            
            # Export options
            st.markdown("#### Export Options")
            
            export_cols = st.columns(2)
            
            with export_cols[0]:
                if st.button("Export to JSON"):
                    # Export to JSON
                    json_data = export_prices_to_json()
                    
                    # Create a download link
                    b64 = base64.b64encode(json_data.encode()).decode()
                    href = f'<a href="data:application/json;base64,{b64}" download="price_data_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.json">Download JSON</a>'
                    st.markdown(href, unsafe_allow_html=True)
                    
                    # Display the JSON data
                    st.code(json_data, language="json")
            
            with export_cols[1]:
                if st.button("Export to CSV"):
                    # Export to CSV
                    csv_data = export_prices_to_csv()
                    
                    # Create a download link
                    b64 = base64.b64encode(csv_data.encode()).decode()
                    href = f'<a href="data:text/csv;base64,{b64}" download="price_data_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv">Download CSV</a>'
                    st.markdown(href, unsafe_allow_html=True)
                    
                    # Display the CSV data preview
                    st.code(csv_data[:500] + "..." if len(csv_data) > 500 else csv_data, language="text")

# Settings Page
def settings_page():
    st.title("⚙️ Settings")
    
    # Get current settings
    settings = get_settings()
    
    # Create tabs for different settings
    scraper_tab, thresholds_tab, misc_tab = st.tabs(["Scraper Settings", "Price Thresholds", "Miscellaneous"])
    
    with scraper_tab:
        st.markdown("### Scraper Settings")
        st.markdown("Configure automatic scraping of product prices.")
        
        # Get current scraper status
        scheduler_status = get_scheduler_status()
        
        # Display current status
        status_cols = st.columns(3)
        
        with status_cols[0]:
            status_text = "Active" if scheduler_status["running"] else "Inactive"
            st.metric("Scraper Status", status_text)
        
        with status_cols[1]:
            interval_minutes = scheduler_status.get("interval_minutes", 0)
            interval_hours = interval_minutes / 60
            st.metric("Scrape Interval", f"{interval_hours:.1f} hours")
        
        with status_cols[2]:
            last_run = scheduler_status.get("last_run", "Never")
            if last_run == "Never":
                st.metric("Last Run", "Never")
            else:
                st.metric("Last Run", last_run)
        
        # Settings form
        with st.form("scraper_settings_form"):
            # Scraping interval
            interval_hours = st.slider(
                "Scraping Interval (hours)",
                min_value=1.0,
                max_value=72.0,
                value=float(interval_minutes / 60) if interval_minutes > 0 else 12.0,
                step=1.0,
                help="How often to scrape product prices automatically"
            )
            
            # Convert to minutes
            interval_minutes = int(interval_hours * 60)
            
            # Submit button
            submit = st.form_submit_button("Save Settings")
            
            if submit:
                # Update settings
                update_settings(scraping_interval=interval_minutes)
                st.success("Scraper settings updated!")
                
                # Restart the scheduler with the new interval
                stop_scheduler()
                start_scheduler()
                st.info("Scheduler restarted with the new interval.")
        
        # Manual controls
        st.markdown("### Manual Controls")
        
        control_cols = st.columns(3)
        
        with control_cols[0]:
            if st.button("Start Scheduler"):
                if scheduler_status["running"]:
                    st.warning("Scheduler is already running.")
                else:
                    start_scheduler()
                    st.success("Scheduler started!")
                    # Refresh the page
                    st.rerun()
        
        with control_cols[1]:
            if st.button("Stop Scheduler"):
                if not scheduler_status["running"]:
                    st.warning("Scheduler is not running.")
                else:
                    stop_scheduler()
                    st.success("Scheduler stopped!")
                    # Refresh the page
                    st.rerun()
        
        with control_cols[2]:
            if st.button("Run Now"):
                with st.spinner("Running scraper..."):
                    results = run_scraper_now()
                    st.success(f"Scraping completed: {results.get('scraped', 0)} products scraped, {results.get('errors', 0)} errors")
    
    with thresholds_tab:
        st.markdown("### Price Thresholds")
        st.markdown("Configure global price thresholds for AI suggestions.")
        
        # Get current threshold settings
        global_min_threshold = float(settings.get("global_min_price_threshold", -5.0))
        global_max_threshold = float(settings.get("global_max_price_threshold", 15.0))
        
        with st.form("threshold_settings_form"):
            threshold_cols = st.columns(2)
            
            with threshold_cols[0]:
                min_threshold = st.number_input(
                    "Global Min Price Threshold (€)",
                    min_value=-1000.0,
                    max_value=0.0,
                    value=global_min_threshold,
                    step=1.0,
                    help="Minimum allowed price variation in EUR (usually negative)"
                )
            
            with threshold_cols[1]:
                max_threshold = st.number_input(
                    "Global Max Price Threshold (€)",
                    min_value=0.0,
                    max_value=1000.0,
                    value=global_max_threshold,
                    step=1.0,
                    help="Maximum allowed price variation in EUR (usually positive)"
                )
            
            # Submit button
            submit = st.form_submit_button("Save Thresholds")
            
            if submit:
                # Update settings
                update_settings(
                    global_min_price_threshold=min_threshold,
                    global_max_price_threshold=max_threshold
                )
                st.success("Price thresholds updated!")
        
        # Explanation
        st.markdown("""
        #### About Price Thresholds
        
        Price thresholds define the allowed range for price suggestions in absolute EUR amounts:
        
        - **Min Threshold (€)**: The minimum amount below the current price a suggestion can go
        - **Max Threshold (€)**: The maximum amount above the current price a suggestion can go
        
        For example, with a current price of €100:
        - A Min Threshold of -€5 means the lowest suggested price would be €95
        - A Max Threshold of €15 means the highest suggested price would be €115
        
        You can override these global settings for individual products.
        """)
    
    with misc_tab:
        st.markdown("### Miscellaneous Settings")
        
        # Analysis period
        analysis_period = int(settings.get("analysis_period", 7))
        
        with st.form("misc_settings_form"):
            new_analysis_period = st.slider(
                "Default Analysis Period (days)",
                min_value=1,
                max_value=60,
                value=analysis_period,
                step=1,
                help="Default number of days to include in price analysis"
            )
            
            # Submit button
            submit = st.form_submit_button("Save Settings")
            
            if submit:
                # Update settings
                update_settings(analysis_period=new_analysis_period)
                st.success("Settings updated!")

# Multi-product Analysis Page
def multi_product_analysis_page():
    st.title("📊 Multi-Product Analysis")
    st.markdown("""
    Analyze and compare multiple products at once. Get AI-powered insights on pricing trends 
    across your product catalog and discover opportunities to optimize your pricing strategy.
    """)
    
    # Get all products
    products_df = get_products()
    
    if products_df.empty:
        st.warning("No products found. Please add products first.")
        return
    
    # Selection mode options
    selection_mode = st.radio(
        "Analysis Mode", 
        ["Analyze All Products", "Select Specific Products"],
        horizontal=True
    )
    
    selected_product_ids = []
    
    if selection_mode == "Analyze All Products":
        selected_product_ids = products_df['id'].tolist()
        st.success(f"Analyzing all {len(selected_product_ids)} products")
    else:
        # Create a multiselect for choosing products
        product_options = [(row['id'], row['name']) for _, row in products_df.iterrows()]
        selected_options = st.multiselect(
            "Select Products to Analyze",
            options=product_options,
            format_func=lambda x: f"{x[1]}"
        )
        
        if not selected_options:
            st.warning("Please select at least one product to analyze.")
            return
        
        selected_product_ids = [option[0] for option in selected_options]
        st.success(f"Analyzing {len(selected_product_ids)} selected products")
    
    # Time period selection
    st.subheader("Analysis Settings")
    col1, col2 = st.columns(2)
    
    with col1:
        time_options = {
            "Last 3 Days": 3,
            "Last 7 Days": 7,
            "Last 14 Days": 14,
            "Last 30 Days": 30,
            "All Time": None
        }
        
        selected_time = st.select_slider(
            "Time Period",
            options=list(time_options.keys()),
            value="Last 7 Days"
        )
        
        days = time_options[selected_time]
    
    with col2:
        analyze_button = st.button("Analyze Products", type="primary")
    
    if analyze_button:
        with st.spinner("Analyzing products..."):
            st.subheader("AI Analysis Report")
            
            # Perform bulk analysis
            if len(selected_product_ids) > 1:
                # For multiple products, use bulk analysis
                analysis_results = get_bulk_analysis(days)
                
                # Filter for only selected products
                analysis_results = [r for r in analysis_results if r['product_id'] in selected_product_ids]
                
                if not analysis_results:
                    st.warning("No analysis data available for the selected products and time period.")
                    return
                
                # Extract key insights
                insights_tab, price_comp_tab, trends_tab, data_tab = st.tabs([
                    "Key Insights", "Price Comparison", "Price Trends", "Raw Data"
                ])
                
                with insights_tab:
                    # Create a summary of insights for all products
                    summary_cols = st.columns(3)
                    
                    # Calculate aggregate metrics
                    total_products = len(analysis_results)
                    products_with_price_changes = sum(1 for r in analysis_results if abs(r.get('price_change_percentage', 0)) > 0.5)
                    avg_price_change = np.mean([r.get('price_change_percentage', 0) for r in analysis_results])
                    products_above_market = sum(1 for r in analysis_results if r.get('price_position', '') == 'above market')
                    products_below_market = sum(1 for r in analysis_results if r.get('price_position', '') == 'below market')
                    products_at_market = total_products - products_above_market - products_below_market
                    
                    with summary_cols[0]:
                        st.metric("Products Analyzed", total_products)
                        st.metric("Products with Price Changes", products_with_price_changes)
                    
                    with summary_cols[1]:
                        st.metric("Average Price Change", f"{avg_price_change:.2f}%", 
                                 delta=f"{avg_price_change:.2f}%" if abs(avg_price_change) > 0.5 else None)
                        
                    with summary_cols[2]:
                        st.metric("Products Above Market", products_above_market)
                        st.metric("Products Below Market", products_below_market)
                    
                    # Show AI insights from bulk analysis
                    st.subheader("AI Insights")
                    
                    for result in analysis_results:
                        with st.expander(f"{result.get('product_name', 'Product')}", expanded=False):
                            # Show AI recommendations
                            if 'recommendation' in result:
                                st.markdown(f"**AI Recommendation:** {result['recommendation']}")
                            
                            # Show reasoning
                            if 'reasoning' in result:
                                st.markdown(f"**Analysis:** {result['reasoning']}")
                            
                            # Show any tips
                            if 'tips' in result:
                                st.markdown(f"**Tips:** {result['tips']}")
                
                with price_comp_tab:
                    # Create a comparative price positions chart
                    pos_data = []
                    for result in analysis_results:
                        # Extract data
                        product_name = result.get('product_name', 'Unknown')
                        our_price = result.get('current_price', 0)
                        avg_competitor = result.get('average_competitor_price', 0)
                        min_competitor = result.get('lowest_competitor_price', 0)
                        max_competitor = result.get('highest_competitor_price', 0)
                        
                        # Add to data list
                        pos_data.append({
                            'product': product_name,
                            'our_price': our_price,
                            'average_competitor': avg_competitor,
                            'lowest_competitor': min_competitor,
                            'highest_competitor': max_competitor,
                            'price_difference_pct': ((our_price - avg_competitor) / avg_competitor * 100) if avg_competitor else 0
                        })
                    
                    pos_df = pd.DataFrame(pos_data)
                    
                    if not pos_df.empty:
                        # Sort by price difference
                        pos_df = pos_df.sort_values('price_difference_pct')
                        
                        # Create the chart
                        fig = go.Figure()
                        
                        # Add trace for our price
                        fig.add_trace(go.Bar(
                            y=pos_df['product'],
                            x=pos_df['our_price'],
                            name='Our Price',
                            orientation='h',
                            marker=dict(color='rgba(58, 71, 80, 0.8)')
                        ))
                        
                        # Add range for competitor prices
                        for i, row in pos_df.iterrows():
                            fig.add_shape(
                                type="line",
                                x0=row['lowest_competitor'], 
                                y0=i,
                                x1=row['highest_competitor'],
                                y1=i,
                                line=dict(color="rgba(156, 165, 196, 1)", width=4),
                                name="Competitor Range"
                            )
                            
                            # Add marker for average
                            fig.add_trace(go.Scatter(
                                x=[row['average_competitor']],
                                y=[row['product']],
                                mode='markers',
                                marker=dict(symbol='diamond', size=10, color='rgba(255, 0, 0, 0.7)'),
                                name='Avg Competitor' if i == 0 else None,
                                showlegend=(i == 0)
                            ))
                        
                        # Update layout
                        fig.update_layout(
                            title='Our Prices vs Competitor Ranges',
                            xaxis_title='Price (€)',
                            yaxis_title='Product',
                            barmode='group',
                            height=max(400, len(pos_df) * 60),
                            legend=dict(
                                orientation="h",
                                yanchor="bottom",
                                y=1.02,
                                xanchor="right",
                                x=1
                            ),
                            margin=dict(l=20, r=20, t=50, b=50),
                        )
                        
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Add explanation
                        st.info("""
                        **Chart explanation:**
                        - **Blue bars** show our current prices
                        - **Gray lines** represent the range of competitor prices (min to max)
                        - **Red diamonds** show the average competitor price
                        """)
                
                with trends_tab:
                    # Create price trend mini-charts for each product
                    st.subheader("Price Trends by Product")
                    
                    # Create a grid of small charts
                    chart_cols = st.columns(2)
                    
                    for i, product_id in enumerate(selected_product_ids):
                        # Get product name
                        product_name = products_df[products_df['id'] == product_id]['name'].iloc[0]
                        
                        # Get price history
                        price_history = get_price_history(product_id, days)
                        
                        if not price_history.empty:
                            with chart_cols[i % 2]:
                                st.markdown(f"#### {product_name}")
                                fig = create_price_history_chart(price_history, product_name, view_mode="line")
                                # Make chart smaller
                                fig.update_layout(height=300, margin=dict(l=20, r=20, t=30, b=20))
                                st.plotly_chart(fig, use_container_width=True)
                
                with data_tab:
                    # Show raw data table
                    st.subheader("Raw Analysis Data")
                    
                    # Create a clean table of results
                    table_data = []
                    for result in analysis_results:
                        row = {
                            'Product': result.get('product_name', 'Unknown'),
                            'Our Price': f"€{result.get('current_price', 0):.2f}",
                            'Avg Competitor': f"€{result.get('average_competitor_price', 0):.2f}",
                            'Price Change': f"{result.get('price_change_percentage', 0):.2f}%",
                            'Position': result.get('price_position', 'Unknown'),
                            'Recommendation': result.get('short_recommendation', 'No recommendation')
                        }
                        table_data.append(row)
                    
                    table_df = pd.DataFrame(table_data)
                    st.dataframe(table_df, use_container_width=True)
            
            else:
                # For a single product, display detailed analysis
                product_id = selected_product_ids[0]
                product_name = products_df[products_df['id'] == product_id]['name'].iloc[0]
                
                st.subheader(f"Detailed Analysis for {product_name}")
                
                # Get single product analysis
                analysis = get_price_analysis(product_id, days)
                
                if not analysis:
                    st.warning("No analysis data available for this product and time period.")
                    return
                
                # Create tabs for different analysis views
                analysis_tab, viz_tab, data_tab = st.tabs([
                    "AI Analysis", "Visualizations", "Raw Data"
                ])
                
                with analysis_tab:
                    # Show current status metrics
                    metric_cols = st.columns(3)
                    
                    with metric_cols[0]:
                        st.metric("Current Price", f"€{analysis.get('current_price', 0):.2f}")
                        
                    with metric_cols[1]:
                        price_change = analysis.get('price_change_percentage', 0)
                        st.metric("Price Change", f"{price_change:.2f}%", 
                                 delta=f"{price_change:.2f}%" if abs(price_change) > 0.5 else None)
                        
                    with metric_cols[2]:
                        st.metric("Market Position", analysis.get('price_position', 'Unknown').title())
                    
                    # Show AI recommendation
                    st.markdown("### AI Recommendation")
                    st.info(analysis.get('recommendation', 'No recommendation available.'))
                    
                    # Show reasoning
                    st.markdown("### Analysis")
                    st.write(analysis.get('reasoning', 'No analysis available.'))
                    
                    # Show tips
                    if 'tips' in analysis:
                        st.markdown("### Tips")
                        st.write(analysis['tips'])
                
                with viz_tab:
                    # Get price history for visualization
                    price_history = get_price_history(product_id, days)
                    
                    if not price_history.empty:
                        # Price trend chart
                        st.markdown("### Price Trend")
                        fig = create_price_history_chart(price_history, product_name, view_mode="line")
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Price comparison with competitors
                        st.markdown("### Competitor Comparison")
                        
                        # Extract competitor data for the latest date
                        latest_data = price_history.iloc[-1]
                        our_price = latest_data['our_price']
                        competitor_prices = latest_data['competitor_prices']
                        
                        # Create comparison chart
                        if competitor_prices and isinstance(competitor_prices, dict):
                            comp_data = {
                                'Seller': ['Our Price'] + list(competitor_prices.keys()),
                                'Price': [our_price] + list(competitor_prices.values())
                            }
                            comp_df = pd.DataFrame(comp_data)
                            
                            # Sort by price
                            comp_df = comp_df.sort_values('Price')
                            
                            # Create bar chart
                            fig = px.bar(
                                comp_df, 
                                x='Seller', 
                                y='Price',
                                color='Seller',
                                color_discrete_map={'Our Price': 'rgba(58, 71, 80, 0.8)'},
                                title="Current Price Comparison"
                            )
                            
                            # Customize
                            fig.update_layout(height=400)
                            st.plotly_chart(fig, use_container_width=True)
                
                with data_tab:
                    # Show raw analysis data
                    st.json(analysis)
    
    # Add information about the feature
    with st.expander("About Multi-Product Analysis", expanded=False):
        st.markdown("""
        ### How to use this feature
        
        The Multi-Product Analysis tool allows you to analyze either all your products or a selected 
        subset to identify pricing trends and opportunities.
        
        **Analysis options:**
        - **Analyze All Products**: Get a comprehensive view of your entire product catalog
        - **Select Specific Products**: Choose specific products to compare and analyze together
        
        **What you get:**
        - AI-powered insights and recommendations for each product
        - Comparative price position analysis across products
        - Price trend visualization for each product
        - Market position assessment (above market, at market, below market)
        
        **Tips:**
        - Use the "Last 7 Days" or "Last 14 Days" timeframes for the most relevant analysis
        - Compare products in the same category to identify category-specific pricing trends
        - Use this analysis before making bulk price adjustments
        """)